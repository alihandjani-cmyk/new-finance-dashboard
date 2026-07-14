#!/usr/bin/env python3
"""
Finance Dashboard Updater
=========================
Fetches data for all 7 exchanges and writes dashboard_data.json.
The HTML reads this file on load — no HTML patching required.

Usage:
  python updater.py          # Full run — news, vol, ILLIQ, spreads, rankings, commentary
  python updater.py --fast   # Fast run — ticker bar, top-10 gainers, market cap only

Scheduled via two GitHub Actions workflows:
  update.yml       — Full run 4× daily (07:30, 10:30, 16:30, 22:00 UTC Mon-Fri)
  update-fast.yml  — Fast run every 30 min during market hours (08:00–21:30 UTC Mon-Fri)

Date contract
-------------
All dates in dashboard_data.json are ISO (YYYY-MM-DD).  Plain string sort is
chronological for ISO — no numeric key needed.  The HTML converts ISO → display
label (e.g. '13 Jul') via fmtDay() at render time.

Universe contract
-----------------
Exchange universes are NOT hardcoded.  On the first full run and then monthly,
build_universe() fetches index constituents from Wikipedia and applies a
$10 M average-daily-value filter.  Results are stored under dash['universes'].
All indicator functions (amihud, spread, AR spread, vol, gainers, market cap)
read from dash['universes'][ex_key] — never from the fallback constants below.
The fallback constants (FTSE_TICKERS etc.) are used ONLY if dash['universes']
is empty (e.g. very first ever run with no internet access to Wikipedia).

Methodology version
-------------------
methodology_version = 'v2_dynamic_universe'
When this key is absent or different, all metric histories are cleared so the
new universe-based series starts clean.

State contract
--------------
Operational state (lse_file_num, universe_refresh_date, commentary dedupe)
lives under the '_state' key of dashboard_data.json, which IS committed by
the workflows.  CI runners are ephemeral — never rely on local files for state.

Requirements (requirements.txt):
    openpyxl>=3.1  yfinance>=0.2  pandas>=2.0  anthropic>=0.25
    requests>=2.31  certifi>=2024.2  lxml>=4.9
"""

import argparse, io, json, math, os, re, ssl, sys, urllib.request, urllib.error, urllib.parse
from datetime import datetime, date, timedelta, timezone
from pathlib import Path

# ─── Fail fast on missing dependencies ───────────────────────────────────────
try:
    import certifi
except ImportError:
    sys.exit('ERROR: certifi not installed. Run: pip install -r requirements.txt')
try:
    import yfinance as yf
except ImportError:
    sys.exit('ERROR: yfinance not installed. Run: pip install -r requirements.txt')
try:
    import pandas as pd
except ImportError:
    sys.exit('ERROR: pandas not installed. Run: pip install -r requirements.txt')
try:
    import openpyxl
except ImportError:
    sys.exit('ERROR: openpyxl not installed. Run: pip install -r requirements.txt')
try:
    import lxml  # noqa: F401 — required by pandas.read_html
except ImportError:
    sys.exit('ERROR: lxml not installed. Run: pip install -r requirements.txt')
try:
    import anthropic as _ant
except ImportError:
    _ant = None  # AI commentary disabled if not installed

# ─── SSL / HTTP ───────────────────────────────────────────────────────────────
_SSL = ssl.create_default_context(cafile=certifi.where())
_HDR = {'User-Agent': 'Mozilla/5.0 (compatible; FinancePulse/1.0)'}

def _get(url, xlsx=True):
    """HTTP GET → bytes, or None on failure."""
    try:
        req = urllib.request.Request(url, headers=_HDR)
        with urllib.request.urlopen(req, context=_SSL, timeout=25) as r:
            if r.status == 200:
                data = r.read()
                if xlsx and data[:2] != b'PK':
                    print(f'  ⚠  non-xlsx response from {url.split("/")[-1]}')
                    return None
                print(f'  ✓  {url.split("?")[0].split("/")[-1]} ({len(data)//1024} KB)')
                return data
    except Exception as exc:
        print(f'  ✗  {url.split("?")[0].split("/")[-1]}: {type(exc).__name__}: {exc}')
    return None

# ─── Paths ────────────────────────────────────────────────────────────────────
_ROOT      = Path(__file__).parent
DATA_FILE  = _ROOT / 'dashboard_data.json'
STATE_FILE = _ROOT / 'updater_state.json'  # gitignored — legacy, no longer written
DAYS       = 90    # rolling history window (business days)

METHODOLOGY_VERSION = 'v2_dynamic_universe'

# ══════════════════════════════════════════════════════════════════════════════
#  EXCHANGE CONFIG — metadata only (no hardcoded tickers)
# ══════════════════════════════════════════════════════════════════════════════

EXCHANGES = {
    'lse':   {'name':'London Stock Exchange','currency':'GBP','vol_currency':'GBP B',
               'vol_method':'file','pence':True, 'nok_eur':False},
    'enx':   {'name':'Euronext',            'currency':'EUR','vol_currency':'EUR B',
               'vol_method':'file','pence':False,'nok_eur':True},
    'ndx':   {'name':'Nasdaq 100',          'currency':'USD','vol_currency':'USD B',
               'vol_method':'yf',  'pence':False,'nok_eur':False},
    'nyse':  {'name':'NYSE',                'currency':'USD','vol_currency':'USD B',
               'vol_method':'yf',  'pence':False,'nok_eur':False},
    'xetra': {'name':'Xetra',               'currency':'EUR','vol_currency':'EUR B',
               'vol_method':'yf',  'pence':False,'nok_eur':False},
    'six':   {'name':'SIX Swiss Exchange',  'currency':'CHF','vol_currency':'CHF B',
               'vol_method':'yf',  'pence':False,'nok_eur':False},
    'tse':   {'name':'Tokyo Stock Exchange','currency':'JPY','vol_currency':'JPY B',
               'vol_method':'yf',  'pence':False,'nok_eur':False},
}

# ── Ticker-bar symbols (live prices shown in the top bar) ────────────────────
TICKER_SYMBOLS = {
    'FTSE':  '^FTSE',    'DAX':    '^GDAXI',   'CAC40': '^FCHI',   'SMI': '^SSMI',
    'SP500': '^GSPC',    'NDX':    '^NDX',      'DJI':   '^DJI',   'NIKKEI': '^N225',
    'GBPUSD':'GBPUSD=X', 'EURUSD': 'EURUSD=X', 'EURGBP':'EURGBP=X','USDJPY':'JPY=X',
    'GOLD':  'GC=F',     'OIL':    'CL=F',
}

# ══════════════════════════════════════════════════════════════════════════════
#  FALLBACK TICKER LISTS  (used ONLY when dash['universes'] is empty)
#  These are NOT the primary data source — build_universe() is.
# ══════════════════════════════════════════════════════════════════════════════

_FALLBACK_TICKERS = {
    'lse': [
        'SHEL.L','AZN.L','HSBA.L','ULVR.L','BP.L','RIO.L','GSK.L','DGE.L','BATS.L','LSEG.L',
        'VOD.L','LLOY.L','NWG.L','PRU.L','NG.L','REL.L','EXPN.L','WPP.L','IMB.L','STAN.L',
        'BARC.L','AAL.L','CNA.L','IHG.L','JD.L','MKS.L','TSCO.L','SBRY.L','RKT.L','HLN.L',
        'ABF.L','ANTO.L','AUTO.L','RR.L','BA.L','CPG.L','BT-A.L','SSE.L','LAND.L','SGRO.L',
        'FLTR.L','MNDI.L','FRES.L','OCDO.L','PSON.L','WEIR.L','IMI.L','GLEN.L',
        'BHP.L','BNZL.L','III.L','SMWH.L','KGF.L','INF.L','ADM.L','LGEN.L','BTRW.L','TW.L',
    ],
    'enx': [
        'MC.PA','OR.PA','RMS.PA','SAN.PA','SU.PA','TTE.PA','AIR.PA','AI.PA','SAF.PA','BNP.PA',
        'KER.PA','DSY.PA','DG.PA','AXA.PA','ENGI.PA','GLE.PA','ORA.PA','SGO.PA','RI.PA','CAP.PA',
        'STM.PA','HO.PA','VIE.PA','RNO.PA','ML.PA','EL.PA','PUB.PA','SW.PA',
        'ASML.AS','INGA.AS','PHIA.AS','ADYEN.AS','UNA.AS','WKL.AS','RAND.AS','ABN.AS','NN.AS','HEIA.AS',
        'ENEL.MI','ENI.MI','ISP.MI','UCG.MI','G.MI','LDO.MI','PRY.MI',
        'UCB.BR','KBC.BR','ABI.BR','EQNR.OL','DNB.OL',
    ],
    'ndx': [
        'AAPL','MSFT','NVDA','AMZN','META','GOOGL','TSLA','AVGO','COST',
        'NFLX','AMD','TMUS','QCOM','INTU','AMAT','BKNG','ISRG','AMGN','CMCSA',
        'TXN','HON','PEP','VRTX','SBUX','GILD','MDLZ','ADI','REGN','KDP',
        'PANW','KLAC','MELI','LRCX','SNPS','CDNS','CTAS','MU','MAR','ORLY',
        'CRWD','CSX','MRVL','PCAR','WDAY','AEP','FTNT','ABNB','MNST','PYPL',
        'CHTR','ODFL','DXCM','EXC','FAST','CEG','ROST','VRSK','CPRT','IDXX',
        'PAYX','TTD','KHC','BIIB','XEL','ZS','ANSS','ON','TEAM','DDOG',
        'WBD','DLTR','CDW','ILMN','MCHP','NXPI','ADSK','GEHC','INTC','CSCO',
        'ADBE','APP','SMCI','ARM','LIN','ANET','FANG','EBAY','PLTR','ASML',
    ],
    'nyse': [
        'BRK-B','JPM','V','JNJ','WMT','PG','XOM','MA','CVX','BAC',
        'LLY','UNH','KO','MRK','ABBV','DIS','GS','MS','CAT','IBM',
        'AXP','DE','C','WFC','RTX','LMT','PFE','BA','T','VZ',
        'COP','GE','NKE','MCD','ACN','PM','CRM','MMM','BMY','NEE',
        'AMT','TGT','HD','LOW','SYK','CI','CB','PLD','SO','DUK',
    ],
    'xetra': [
        'SAP.DE','SIE.DE','ALV.DE','DTE.DE','BMW.DE','MBG.DE','BAS.DE','MUV2.DE',
        'ADS.DE','RWE.DE','DBK.DE','BAYN.DE','IFX.DE','EOAN.DE','VOW3.DE','MTX.DE',
        'DTG.DE','VNA.DE','DB1.DE','RHM.DE','P911.DE','BEI.DE','HEI.DE','BNR.DE',
        'ENR.DE','DHL.DE','CBK.DE','PAH3.DE','MRK.DE','HNR1.DE','PUM.DE','SRT3.DE',
        'ZAL.DE','FME.DE','QIA.DE','SY1.DE','CON.DE','HLAG.DE',
    ],
    'six': [
        'NESN.SW','NOVN.SW','ROG.SW','UBSG.SW','ALC.SW','ABBN.SW','SREN.SW',
        'GEBN.SW','GIVN.SW','LOGN.SW','PGHN.SW','SCMN.SW','SIKA.SW','SLHN.SW',
        'SOON.SW','CFR.SW','ZURN.SW','HOLN.SW','STMN.SW','VACN.SW','KNIN.SW',
        'LISN.SW','BALN.SW',
    ],
    'tse': [
        '7203.T','6758.T','8306.T','9984.T','6861.T','7974.T','9983.T','9432.T',
        '9433.T','4502.T','7267.T','7201.T','7751.T','6752.T','6501.T','8058.T',
        '8031.T','8001.T','8316.T','8411.T','8035.T','6857.T','6098.T','4568.T',
        '4519.T','4063.T','6954.T','6981.T','6594.T','6902.T','5108.T','4452.T',
        '9020.T','9022.T','9202.T','8604.T','8766.T','3382.T','4755.T','4911.T',
        '6301.T','6506.T','6326.T','5401.T','8802.T','8801.T','9531.T','6367.T',
        '4523.T','6273.T',
    ],
}

_FALLBACK_NAMES = {
    'lse': {
        'AZN.L':'AstraZeneca','SHEL.L':'Shell','HSBA.L':'HSBC Holdings','ULVR.L':'Unilever',
        'BP.L':'BP','RIO.L':'Rio Tinto','GSK.L':'GSK','DGE.L':'Diageo',
        'BATS.L':'British American Tobacco','LSEG.L':'London Stock Exchange Group',
        'VOD.L':'Vodafone','LLOY.L':'Lloyds Banking Group','NWG.L':'NatWest Group',
        'PRU.L':'Prudential','NG.L':'National Grid','REL.L':'RELX','EXPN.L':'Experian',
        'WPP.L':'WPP','IMB.L':'Imperial Brands','STAN.L':'Standard Chartered',
        'BARC.L':'Barclays','AAL.L':'Anglo American','CNA.L':'Centrica',
        'IHG.L':'IHG Hotels & Resorts','JD.L':'JD Sports Fashion',
        'MKS.L':'Marks & Spencer','TSCO.L':'Tesco','SBRY.L':"Sainsbury's",
        'RKT.L':'Reckitt','HLN.L':'Haleon','ABF.L':'Associated British Foods',
        'ANTO.L':'Antofagasta','AUTO.L':'Auto Trader','RR.L':'Rolls-Royce Holdings',
        'BA.L':'BAE Systems','CPG.L':'Compass Group','BT-A.L':'BT Group','SSE.L':'SSE',
        'LAND.L':'Land Securities','SGRO.L':'Segro','FLTR.L':'Flutter Entertainment',
        'MNDI.L':'Mondi','FRES.L':'Fresnillo','OCDO.L':'Ocado Group','PSON.L':'Pearson',
        'WEIR.L':'Weir Group','IMI.L':'IMI','GLEN.L':'Glencore','BHP.L':'BHP Group',
        'BNZL.L':'Bunzl','III.L':'3i Group','SMWH.L':'WH Smith','KGF.L':'Kingfisher',
        'INF.L':'Informa','ADM.L':'Admiral Group','LGEN.L':'Legal & General',
        'BTRW.L':'Barratt Redrow','TW.L':'Taylor Wimpey',
    },
    'enx': {
        'MC.PA':'LVMH','OR.PA':"L'Oréal",'RMS.PA':'Hermès','SAN.PA':'Sanofi',
        'SU.PA':'Schneider Electric','TTE.PA':'TotalEnergies','AIR.PA':'Airbus',
        'AI.PA':'Air Liquide','SAF.PA':'Safran','BNP.PA':'BNP Paribas','KER.PA':'Kering',
        'DSY.PA':'Dassault Systèmes','DG.PA':'Vinci','AXA.PA':'AXA','ENGI.PA':'Engie',
        'GLE.PA':'Société Générale','ORA.PA':'Orange','SGO.PA':'Saint-Gobain',
        'RI.PA':'Pernod Ricard','CAP.PA':'Capgemini','STM.PA':'STMicroelectronics',
        'HO.PA':'Thales','VIE.PA':'Veolia','RNO.PA':'Renault','ML.PA':'Michelin',
        'EL.PA':'EssilorLuxottica','PUB.PA':'Publicis','SW.PA':'Sodexo',
        'ASML.AS':'ASML Holding','INGA.AS':'ING Group','PHIA.AS':'Philips',
        'ADYEN.AS':'Adyen','UNA.AS':'Unilever NV','WKL.AS':'Wolters Kluwer',
        'RAND.AS':'Randstad','ABN.AS':'ABN AMRO','NN.AS':'NN Group','HEIA.AS':'Heineken',
        'ENEL.MI':'Enel','ENI.MI':'ENI','ISP.MI':'Intesa Sanpaolo','UCG.MI':'UniCredit',
        'G.MI':'Generali','LDO.MI':'Leonardo','PRY.MI':'Prysmian',
        'UCB.BR':'UCB','KBC.BR':'KBC Group','ABI.BR':'AB InBev',
        'EQNR.OL':'Equinor','DNB.OL':'DNB Bank',
    },
    'ndx': {
        'AAPL':'Apple','MSFT':'Microsoft','NVDA':'NVIDIA','AMZN':'Amazon',
        'META':'Meta Platforms','GOOGL':'Alphabet','TSLA':'Tesla','AVGO':'Broadcom',
        'COST':'Costco','NFLX':'Netflix','AMD':'AMD','TMUS':'T-Mobile US',
        'QCOM':'Qualcomm','INTU':'Intuit','AMAT':'Applied Materials',
        'BKNG':'Booking Holdings','ISRG':'Intuitive Surgical','AMGN':'Amgen',
        'CMCSA':'Comcast','TXN':'Texas Instruments','HON':'Honeywell','PEP':'PepsiCo',
        'VRTX':'Vertex Pharma','SBUX':'Starbucks','GILD':'Gilead Sciences',
        'MDLZ':'Mondelēz Intl','ADI':'Analog Devices','REGN':'Regeneron',
        'KDP':'Keurig Dr Pepper','PANW':'Palo Alto Networks','KLAC':'KLA Corp',
        'MELI':'MercadoLibre','LRCX':'Lam Research','SNPS':'Synopsys',
        'CDNS':'Cadence Design','CTAS':'Cintas','MU':'Micron Technology',
        'MAR':'Marriott','ORLY':"O'Reilly Auto",'CRWD':'CrowdStrike','CSX':'CSX Corp',
        'MRVL':'Marvell Technology','PCAR':'PACCAR','WDAY':'Workday',
        'AEP':'American Electric','FTNT':'Fortinet','ABNB':'Airbnb',
        'MNST':'Monster Beverage','PYPL':'PayPal','CHTR':'Charter Comms',
        'ODFL':'Old Dominion Freight','DXCM':'DexCom','EXC':'Exelon','FAST':'Fastenal',
        'CEG':'Constellation Energy','ROST':'Ross Stores','VRSK':'Verisk Analytics',
        'CPRT':'Copart','IDXX':'IDEXX Labs','PAYX':'Paychex','TTD':'The Trade Desk',
        'KHC':'Kraft Heinz','BIIB':'Biogen','XEL':'Xcel Energy','ZS':'Zscaler',
        'ANSS':'ANSYS','ON':'ON Semiconductor','TEAM':'Atlassian','DDOG':'Datadog',
        'WBD':'Warner Bros. Discovery','DLTR':'Dollar Tree','CDW':'CDW Corp',
        'ILMN':'Illumina','MCHP':'Microchip Tech','NXPI':'NXP Semiconductors',
        'ADSK':'Autodesk','GEHC':'GE HealthCare','INTC':'Intel','CSCO':'Cisco',
        'ADBE':'Adobe','APP':'AppLovin','SMCI':'Super Micro Computer','ARM':'Arm Holdings',
        'LIN':'Linde','ANET':'Arista Networks','FANG':'Diamondback Energy',
        'EBAY':'eBay','PLTR':'Palantir','ASML':'ASML Holding',
    },
    'nyse': {
        'BRK-B':'Berkshire Hathaway','JPM':'JPMorgan Chase','V':'Visa',
        'JNJ':'Johnson & Johnson','WMT':'Walmart','PG':'Procter & Gamble',
        'XOM':'ExxonMobil','MA':'Mastercard','CVX':'Chevron','BAC':'Bank of America',
        'LLY':'Eli Lilly','UNH':'UnitedHealth','KO':'Coca-Cola','MRK':'Merck',
        'ABBV':'AbbVie','DIS':'Walt Disney','GS':'Goldman Sachs','MS':'Morgan Stanley',
        'CAT':'Caterpillar','IBM':'IBM','AXP':'American Express','DE':'Deere & Co',
        'C':'Citigroup','WFC':'Wells Fargo','RTX':'RTX Corp','LMT':'Lockheed Martin',
        'PFE':'Pfizer','BA':'Boeing','T':'AT&T','VZ':'Verizon',
        'COP':'ConocoPhillips','GE':'GE Aerospace','NKE':'Nike','MCD':"McDonald's",
        'ACN':'Accenture','PM':'Philip Morris Intl','CRM':'Salesforce','MMM':'3M',
        'BMY':'Bristol-Myers Squibb','NEE':'NextEra Energy','AMT':'American Tower',
        'TGT':'Target','HD':'Home Depot','LOW':"Lowe's",'SYK':'Stryker',
        'CI':'Cigna','CB':'Chubb','PLD':'Prologis','SO':'Southern Company','DUK':'Duke Energy',
    },
    'xetra': {
        'SAP.DE':'SAP','SIE.DE':'Siemens','ALV.DE':'Allianz','DTE.DE':'Deutsche Telekom',
        'BMW.DE':'BMW','MBG.DE':'Mercedes-Benz','BAS.DE':'BASF','MUV2.DE':'Munich Re',
        'ADS.DE':'Adidas','RWE.DE':'RWE','DBK.DE':'Deutsche Bank','BAYN.DE':'Bayer',
        'IFX.DE':'Infineon','EOAN.DE':'E.ON','VOW3.DE':'Volkswagen',
        'MTX.DE':'MTU Aero Engines','DTG.DE':'Daimler Truck','VNA.DE':'Vonovia',
        'DB1.DE':'Deutsche Boerse','RHM.DE':'Rheinmetall','P911.DE':'Porsche AG',
        'BEI.DE':'Beiersdorf','HEI.DE':'Heidelberg Materials','BNR.DE':'Brenntag',
        'ENR.DE':'Siemens Energy','DHL.DE':'DHL Group','CBK.DE':'Commerzbank',
        'PAH3.DE':'Porsche Holding','MRK.DE':'Merck KGaA','HNR1.DE':'Hannover Re',
        'PUM.DE':'Puma','SRT3.DE':'Sartorius','ZAL.DE':'Zalando',
        'FME.DE':'Fresenius Medical','QIA.DE':'Qiagen','SY1.DE':'Symrise',
        'CON.DE':'Continental','HLAG.DE':'Hapag-Lloyd',
    },
    'six': {
        'NESN.SW':'Nestlé','NOVN.SW':'Novartis','ROG.SW':'Roche','UBSG.SW':'UBS Group',
        'ALC.SW':'Alcon','ABBN.SW':'ABB','SREN.SW':'Swiss Re','GEBN.SW':'Geberit',
        'GIVN.SW':'Givaudan','LOGN.SW':'Lonza Group','PGHN.SW':'Partners Group',
        'SCMN.SW':'Swisscom','SIKA.SW':'Sika','SLHN.SW':'Swiss Life','SOON.SW':'Sonova',
        'CFR.SW':'Richemont','ZURN.SW':'Zurich Insurance','HOLN.SW':'Holcim',
        'STMN.SW':'Straumann','VACN.SW':'VAT Group','KNIN.SW':'Kuehne+Nagel',
        'LISN.SW':'Lindt & Spruengli','BALN.SW':'Baloise',
    },
    'tse': {
        '7203.T':'Toyota Motor','6758.T':'Sony Group','8306.T':'Mitsubishi UFJ Financial',
        '9984.T':'SoftBank Group','6861.T':'Keyence','7974.T':'Nintendo',
        '9983.T':'Fast Retailing','9432.T':'Nippon Telegraph & Telephone','9433.T':'KDDI',
        '4502.T':'Takeda Pharmaceutical','7267.T':'Honda Motor','7201.T':'Nissan Motor',
        '7751.T':'Canon','6752.T':'Panasonic Holdings','6501.T':'Hitachi',
        '8058.T':'Mitsubishi Corp','8031.T':'Mitsui & Co','8001.T':'Itochu',
        '8316.T':'Sumitomo Mitsui Financial','8411.T':'Mizuho Financial',
        '8035.T':'Tokyo Electron','6857.T':'Advantest','6098.T':'Recruit Holdings',
        '4568.T':'Daiichi Sankyo','4519.T':'Chugai Pharmaceutical',
        '4063.T':'Shin-Etsu Chemical','6954.T':'Fanuc','6981.T':'Murata Manufacturing',
        '6594.T':'Nidec','6902.T':'Denso','5108.T':'Bridgestone','4452.T':'Kao',
        '9020.T':'East Japan Railway','9022.T':'Central Japan Railway',
        '9202.T':'ANA Holdings','8604.T':'Nomura Holdings','8766.T':'Tokio Marine Holdings',
        '3382.T':'Seven & I Holdings','4755.T':'Rakuten Group','4911.T':'Shiseido',
        '6301.T':'Komatsu','6506.T':'Yaskawa Electric','6326.T':'Kubota',
        '5401.T':'Nippon Steel','8802.T':'Mitsubishi Estate','8801.T':'Mitsui Fudosan',
        '9531.T':'Tokyo Gas','6367.T':'Daikin Industries','4523.T':'Eisai',
        '6273.T':'SMC Corp',
    },
}

# ══════════════════════════════════════════════════════════════════════════════
#  DATA FILE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _load_dashboard():
    """Load existing dashboard_data.json (preserves 90-day histories)."""
    try:
        if DATA_FILE.exists():
            return json.loads(DATA_FILE.read_text())
    except Exception as exc:
        print(f'  ⚠  Could not load existing data: {exc}')
    return {
        'generated_at': None, 'stale_fallback': True, 'commentary': None,
        'commentary_date': None, 'methodology_version': None,
        'news': [], 'tickers': {},
        'universes':  {k: {} for k in EXCHANGES},
        'gainers':    {k: [] for k in EXCHANGES},
        'market_cap': {k: {'date': None, 'currency': v['currency'], 'top10': []} for k, v in EXCHANGES.items()},
        'vol':        {k: {'currency': v['vol_currency'], 'dates': [], 'value': []} for k, v in EXCHANGES.items()},
        'vol_comparable': {k: {
            'currency': v['vol_currency'], 'currency_usd': 'USD B',
            'dates': [], 'value': [], 'value_usd': [], 'shares': [], 'n_universe': 0,
        } for k, v in EXCHANGES.items()},
        'amihud':     {k: {'history': [], 'marketAvg': None} for k in EXCHANGES},
        'ar_spread':       {k: {'history': [], 'marketAvg': None} for k in EXCHANGES},
        'turnover_ratio':  {k: {'history': []} for k in EXCHANGES},
        'current_ranking': {'date': None, 'ranks': {k: {
            'vol': None, 'illiq': None, 'tr': None, 'ar': None, 'composite': None,
        } for k in EXCHANGES}},
        'ranking_history': [],
    }

def _save_dashboard(data):
    data['generated_at'] = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    data['stale_fallback'] = False
    DATA_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    print(f'\n  ✓  Saved {DATA_FILE} ({DATA_FILE.stat().st_size // 1024} KB)')

# ─── Date / history utilities ─────────────────────────────────────────────────

def _migrate_dates(dash):
    """One-time migration: convert any remaining 'dd Mon' dates to ISO."""
    mon_map = {'Jan':'01','Feb':'02','Mar':'03','Apr':'04','May':'05','Jun':'06',
               'Jul':'07','Aug':'08','Sep':'09','Oct':'10','Nov':'11','Dec':'12'}
    today = date.today()

    def _to_iso(label):
        if not label or '-' in str(label):
            return label
        try:
            parts = str(label).strip().split()
            if len(parts) != 2:
                return label
            d_str, m_str = parts
            m_num = mon_map.get(m_str)
            if not m_num:
                return label
            d_num = int(d_str)
            year = today.year
            if int(m_num) > today.month or (int(m_num) == today.month and d_num > today.day):
                year -= 1
            return f'{year}-{m_num}-{d_num:02d}'
        except Exception:
            return label

    migrated = 0
    for section in ('amihud', 'spread', 'ar_spread', 'turnover_ratio'):
        for ex in dash.get(section, {}).values():
            for entry in ex.get('history', []):
                if 'date' in entry and '-' not in str(entry['date']):
                    entry['date'] = _to_iso(entry['date'])
                    migrated += 1
    for ex in dash.get('vol', {}).values():
        old_dates = ex.get('dates', [])
        new_dates = [_to_iso(d) for d in old_dates]
        if new_dates != old_dates:
            ex['dates'] = new_dates
            migrated += len(old_dates)
    for entry in dash.get('ranking_history', []):
        if 'date' in entry and '-' not in str(entry['date']):
            entry['date'] = _to_iso(entry['date'])
            migrated += 1
    cr = dash.get('current_ranking', {})
    if cr.get('date') and '-' not in str(cr['date']):
        cr['date'] = _to_iso(cr['date'])
        migrated += 1
    if migrated:
        print(f'  ℹ  Migrated {migrated} legacy date labels to ISO format')
    return dash


def _push(lst, entry, key='date', maxn=DAYS):
    """Append/update entry (matched by key), sort by ISO date string, trim to maxn."""
    out = [e for e in lst if e.get(key) != entry.get(key)]
    out.append(entry)
    out.sort(key=lambda e: e.get(key, ''))
    return out[-maxn:]

def _today():
    return date.today().isoformat()

def _prev_bday(dt=None, n=1):
    d = dt or date.today()
    for _ in range(n):
        d -= timedelta(days=1)
        while d.weekday() >= 5:
            d -= timedelta(days=1)
    return d

def _trim_stale_dates(date_keys, n=5, max_gap_days=4):
    """Return up to n most recent ISO date strings from date_keys, walking
    backward from the newest. Stops early if the gap to the next-older date
    exceeds max_gap_days (covers a normal weekend/holiday gap of ~3 days).
    This prevents an old leftover date from a failed fetch (e.g. a source
    outage) from lingering in a 'last 5' window for days once fewer than n
    fresh dates have accumulated since the gap — better to show fewer, correct
    points than a stale, disconnected one.
    """
    all_sorted = sorted(date_keys)
    if not all_sorted:
        return []
    kept = [all_sorted[-1]]
    for label in reversed(all_sorted[:-1]):
        if len(kept) >= n:
            break
        newest_kept = date.fromisoformat(kept[-1])
        gap = (newest_kept - date.fromisoformat(label)).days
        if gap > max_gap_days:
            break
        kept.append(label)
    return list(reversed(kept))

# ══════════════════════════════════════════════════════════════════════════════
#  UNIVERSE SELECTION — dynamic index-constituent + ADV filter
#  Refreshed monthly; stored under dash['universes'][ex_key].
# ══════════════════════════════════════════════════════════════════════════════

# Wikipedia sources: (url, yfinance_suffix)
# For ENX we use individual national index pages (one suffix per page) rather
# than the Euronext 100 page whose tickers require country-lookup to suffix.
_WIKI_SOURCES = {
    'lse':   [
        ('https://en.wikipedia.org/wiki/FTSE_100',       '.L'),
        ('https://en.wikipedia.org/wiki/FTSE_250_Index', '.L'),
    ],
    'enx':   [
        ('https://en.wikipedia.org/wiki/CAC_40',          '.PA'),
        ('https://en.wikipedia.org/wiki/AEX_index',       '.AS'),
        ('https://en.wikipedia.org/wiki/FTSE_MIB',        '.MI'),
        ('https://en.wikipedia.org/wiki/BEL_20',          '.BR'),
        ('https://en.wikipedia.org/wiki/PSI-20',          '.LS'),
        ('https://en.wikipedia.org/wiki/OBX_Stock_Index', '.OL'),
    ],
    'ndx':   [
        ('https://en.wikipedia.org/wiki/Nasdaq-100', None),
    ],
    'nyse':  [
        # S&P 500 — filtered below to NYSE+NYSE Arca listings only
        ('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies', None),
    ],
    'xetra': [
        ('https://en.wikipedia.org/wiki/DAX',  '.DE'),
        ('https://en.wikipedia.org/wiki/MDAX', '.DE'),
    ],
    'six':   [
        ('https://en.wikipedia.org/wiki/Swiss_Market_Index', '.SW'),
        ('https://en.wikipedia.org/wiki/SMIM',               '.SW'),
    ],
    'tse':   [
        ('https://en.wikipedia.org/wiki/Nikkei_225', '.T'),
    ],
}

# Exchanges whose native ticker codes are numeric (or digit-led alphanumeric,
# e.g. TSE's newer 4-char codes like '543A') rather than letters. _clean_ticker
# rejects digit-led values by default to filter out stray footnote numbers on
# Western index pages — these exchanges opt out of that check.
_NUMERIC_TICKER_EXCHANGES = {'tse'}

# Column name aliases used to locate ticker and name columns in Wikipedia tables
_TICKER_ALIASES = ['ticker','symbol','epic','code','stock ticker','ticker symbol',
                   'trading symbol','exchange symbol','index ticker','abbr.','abbreviation']
_NAME_ALIASES   = ['company','name','security','company name','stock','constituent',
                   'issuer','organisation']

def _find_col(df, aliases):
    """Return the first df column whose lowercased name matches an alias, or None."""
    lc = {str(c).lower().strip(): c for c in df.columns}
    for a in aliases:
        if a in lc:
            return lc[a]
    return None

def _clean_ticker(raw, allow_numeric=False):
    """Strip footnote markers, whitespace; return cleaned ticker string or ''.

    allow_numeric : for exchanges whose native codes are digit-led (e.g. TSE's
    4-character codes: '7203', or newer alphanumeric ones like '543A'). The
    raw cell may carry extra text or wrap the code in parens ('(TYO: 7203)'),
    so this searches for the code pattern directly rather than stripping
    parens first (which would delete the code itself) or requiring the whole
    cleaned cell to equal it.
    """
    if allow_numeric:
        m = re.search(r'\b(\d{3}[0-9A-Z])\b', str(raw).upper())
        return m.group(1) if m else ''
    s = re.sub(r'\[.*?\]|\(.*?\)', '', str(raw)).strip()
    # Reject obviously non-ticker values
    if not s or s in ('nan', 'NaN', '-', '—', 'N/A') or len(s) > 15:
        return ''
    if s[0].isdigit():
        return ''
    return s

def _parse_wiki_table(url, suffix, filter_nyse=False, allow_numeric=False):
    """Fetch a Wikipedia page and extract the constituent ticker→name dict.

    suffix        : string appended to raw ticker if ticker has no '.', or None.
    filter_nyse   : if True, skip rows where Exchange column contains 'nasdaq'.
    allow_numeric : if True, accept digit-led ticker codes (see _clean_ticker).
    Returns {yf_ticker: display_name} or {} on failure.
    """
    raw = _get(url, xlsx=False)
    if not raw:
        return {}
    try:
        html = raw.decode('utf-8', errors='replace')
        tables = pd.read_html(io.StringIO(html))
    except Exception as exc:
        print(f'    ⚠  HTML parse error ({url.split("/")[-1]}): {exc}')
        return {}

    for tbl in tables:
        if len(tbl) < 5:
            continue
        ticker_col = _find_col(tbl, _TICKER_ALIASES)
        name_col   = _find_col(tbl, _NAME_ALIASES)
        if ticker_col is None:
            continue

        # For NYSE: look for an exchange column to skip Nasdaq-listed stocks
        exch_col = None
        if filter_nyse:
            exch_col = _find_col(tbl, ['exchange','listing exchange','traded on'])

        result = {}
        for _, row in tbl.iterrows():
            raw_tk = _clean_ticker(row[ticker_col], allow_numeric=allow_numeric)
            if not raw_tk:
                continue

            # NYSE filter: skip Nasdaq rows
            if filter_nyse and exch_col:
                exch_val = str(row[exch_col]).lower()
                if 'nasdaq' in exch_val:
                    continue

            # Add exchange suffix if not already present
            if suffix and '.' not in raw_tk:
                ticker = raw_tk + suffix
            else:
                ticker = raw_tk

            name = (re.sub(r'\[.*?\]|\(.*?\)', '', str(row[name_col])).strip()
                    if name_col else ticker.split('.')[0])
            if not name or name == 'nan':
                name = ticker.split('.')[0]

            result[ticker] = name

        if len(result) >= 5:
            return result

    print(f'    ⚠  No constituent table found at {url.split("/")[-1]}')
    return {}


def _fetch_constituents(ex_key):
    """Fetch raw index constituent {ticker: name} dict from Wikipedia.
    For NYSE: uses S&P 500 minus Nasdaq 100 (belt-and-suspenders approach).
    """
    sources = _WIKI_SOURCES.get(ex_key, [])
    combined = {}
    allow_numeric = ex_key in _NUMERIC_TICKER_EXCHANGES

    for url, suffix in sources:
        print(f'    Wikipedia: {url.split("/")[-1].replace("%26", "&")}')
        chunk = _parse_wiki_table(url, suffix, filter_nyse=(ex_key == 'nyse'), allow_numeric=allow_numeric)
        if chunk:
            combined.update(chunk)
            print(f'      → {len(chunk)} tickers ({len(combined)} total)')
        else:
            print(f'      → empty/failed')

    # NYSE belt-and-suspenders: also subtract Nasdaq 100 tickers explicitly
    if ex_key == 'nyse' and combined:
        ndx_chunk = _parse_wiki_table(
            'https://en.wikipedia.org/wiki/Nasdaq-100', None, filter_nyse=False)
        ndx_set = set(ndx_chunk.keys())
        before = len(combined)
        combined = {t: n for t, n in combined.items() if t not in ndx_set}
        removed = before - len(combined)
        if removed:
            print(f'    Removed {removed} Nasdaq-listed tickers from NYSE universe')

    return combined


def _fetch_fx_rates():
    """Fetch GBPUSD, EURUSD, CHFUSD, JPYUSD (and NOKEUR) for ADV threshold conversion.
    Returns dict: currency_code → USD rate (e.g. {'GBP': 1.27, 'EUR': 1.08, ...}).
    """
    fx_map  = {'GBP': 'GBPUSD=X', 'EUR': 'EURUSD=X', 'CHF': 'CHFUSD=X', 'NOK': 'NOKUSD=X'}
    # Yahoo quotes JPY as USD/JPY ('JPY=X', ~150), not JPY/USD — invert below.
    inverse_fx_map = {'JPY': 'JPY=X'}
    rates   = {'USD': 1.0}
    syms    = list(fx_map.values()) + list(inverse_fx_map.values())
    try:
        raw = yf.download(syms, period='2d', interval='1d',
                          progress=False, auto_adjust=True, threads=True)
        if not raw.empty:
            close = raw['Close'] if isinstance(raw.columns, pd.MultiIndex) else raw
            for ccy, sym in fx_map.items():
                try:
                    if sym in close.columns:
                        rates[ccy] = float(close[sym].dropna().iloc[-1])
                except Exception:
                    pass
            for ccy, sym in inverse_fx_map.items():
                try:
                    if sym in close.columns:
                        usd_per_unit = float(close[sym].dropna().iloc[-1])
                        if usd_per_unit > 0:
                            rates[ccy] = 1.0 / usd_per_unit
                except Exception:
                    pass
    except Exception as exc:
        print(f'  ⚠  FX rates fetch failed ({exc}); using fallback rates')
    rates.setdefault('GBP', 1.27)
    rates.setdefault('EUR', 1.08)
    rates.setdefault('CHF', 1.11)
    rates.setdefault('NOK', 0.086)
    rates.setdefault('JPY', 0.0067)
    print(f'  FX: GBP={rates["GBP"]:.4f} EUR={rates["EUR"]:.4f} '
          f'CHF={rates["CHF"]:.4f} NOK={rates["NOK"]:.4f} JPY={rates["JPY"]:.5f}')
    return rates


def _apply_adv_filter(tickers_names, ex_key, fx_rates, threshold_usd_m=10.0):
    """Download 25 calendar days of OHLCV; keep stocks with 20-day ADV > threshold.

    threshold_usd_m : minimum average daily value in USD millions (default $10 M).
    Returns filtered {ticker: name} dict.  On download failure returns input unchanged.
    """
    tickers  = list(tickers_names.keys())
    if not tickers:
        return {}

    is_pence = EXCHANGES[ex_key].get('pence', False)
    nok_eur  = EXCHANGES[ex_key].get('nok_eur', False)
    currency = EXCHANGES[ex_key]['currency']
    fx_usd   = fx_rates.get(currency, 1.0)   # local currency → USD

    print(f'    ADV filter: {len(tickers)} tickers...')
    try:
        raw = yf.download(tickers, period='25d', interval='1d',
                          progress=False, auto_adjust=True, threads=True)
        if raw.empty:
            print(f'    ⚠  ADV download empty — returning unfiltered')
            return tickers_names
    except Exception as exc:
        print(f'    ⚠  ADV download failed ({exc}) — returning unfiltered')
        return tickers_names

    is_multi = isinstance(raw.columns, pd.MultiIndex)
    if not is_multi:
        # Single ticker edge case
        vals = raw['Close'] * raw['Volume']
        if is_pence:
            vals = vals / 100
        adv_local = float(vals.tail(20).mean())
        adv_usd_m = adv_local / 1e6 * fx_usd
        return tickers_names if adv_usd_m >= threshold_usd_m else {}

    close_df  = raw['Close']
    volume_df = raw['Volume']
    daily_val = close_df * volume_df   # pence·shares or ccy·shares

    if is_pence:
        daily_val = daily_val / 100    # → £·shares

    if nok_eur:
        # Norwegian Oslo-listed stocks: value is in NOK, convert to EUR then USD
        nok_eur_rate = fx_rates.get('NOK', 0.086) / fx_rates.get('EUR', 1.08)
        ol_cols = [t for t in tickers if t.endswith('.OL') and t in daily_val.columns]
        for t in ol_cols:
            daily_val[t] = daily_val[t] * nok_eur_rate   # NOK → EUR

    # 20-day ADV → USD millions
    adv_local = daily_val.tail(20).mean()        # Series: ticker → avg local value
    adv_usd_m = adv_local * fx_usd / 1e6        # → USD millions

    filtered = {}
    passed = failed = skipped = 0
    for ticker, name in tickers_names.items():
        if ticker not in adv_usd_m.index:
            skipped += 1
            continue
        val = adv_usd_m[ticker]
        if pd.isna(val) or val < threshold_usd_m:
            failed += 1
        else:
            filtered[ticker] = name
            passed += 1

    print(f'    ADV result: {passed} passed / {failed} below threshold / '
          f'{skipped} no data  (threshold ${threshold_usd_m:.0f}M USD)')
    return filtered


def build_universe(ex_key, fx_rates):
    """Build the liquid universe for one exchange.

    1. Fetch index constituents from Wikipedia.
    2. Apply $10 M ADV filter.
    3. Return universe record dict, or None on failure.
    """
    _INDEX_LABEL = {
        'lse':   'FTSE 100 + FTSE 250',
        'enx':   'CAC 40 + AEX + FTSE MIB + BEL 20 + PSI 20 + OBX',
        'ndx':   'Nasdaq 100',
        'nyse':  'S&P 500 (NYSE + NYSE Arca)',
        'xetra': 'DAX 40 + MDAX 60',
        'six':   'SMI 20 + SMIM 30',
        'tse':   'Nikkei 225',
    }
    print(f'  Building universe: {ex_key} ({_INDEX_LABEL.get(ex_key, ex_key)})')

    raw = _fetch_constituents(ex_key)
    if not raw:
        print(f'  ✗  {ex_key}: constituent fetch empty — universe not rebuilt')
        return None

    n_raw = len(raw)
    filtered = _apply_adv_filter(raw, ex_key, fx_rates)
    n_filtered = len(filtered)

    if n_filtered < 5:
        print(f'  ✗  {ex_key}: only {n_filtered} stocks passed ADV filter — universe not rebuilt')
        return None

    print(f'  ✓  {ex_key}: {n_filtered}/{n_raw} stocks in liquid universe')
    return {
        'tickers':             list(filtered.keys()),
        'names':               filtered,
        'index_source':        _INDEX_LABEL.get(ex_key, ex_key),
        'n_constituents':      n_raw,
        'n_universe':          n_filtered,
        'adv_threshold_usd_m': 10,
        'refreshed':           _today(),
    }


def _get_universe(dash, ex_key):
    """Return (tickers, names) from the dynamic universe, falling back to
    hardcoded lists if no universe has been built yet."""
    uni = dash.get('universes', {}).get(ex_key)
    if uni and uni.get('tickers'):
        return uni['tickers'], uni['names']
    # Emergency fallback — used only on very first run before Wikipedia succeeds
    print(f'  ⚠  {ex_key}: no dynamic universe — using fallback hardcoded list')
    tickers = _FALLBACK_TICKERS.get(ex_key, [])
    names   = _FALLBACK_NAMES.get(ex_key, {})
    return tickers, names

# ══════════════════════════════════════════════════════════════════════════════
#  NEWS
# ══════════════════════════════════════════════════════════════════════════════

def fetch_news():
    """Fetch top 10 financial news headlines from multiple RSS feeds."""
    import xml.etree.ElementTree as ET
    feeds = [
        ('https://feeds.bbci.co.uk/news/business/rss.xml', 'BBC Business'),
        ('https://news.google.com/rss/search?q=stock+market+finance&hl=en-US&gl=US&ceid=US:en', None),
        ('https://feeds.marketwatch.com/marketwatch/topstories/', 'MarketWatch'),
        ('https://feeds.cnbc.com/cnbc/ID/100003114/device/rss/rss.html', 'CNBC'),
    ]
    items = []
    for url, default_src in feeds:
        if len(items) >= 10:
            break
        raw = _get(url, xlsx=False)
        if not raw:
            continue
        try:
            root = ET.fromstring(raw)
            for item in root.iter('item'):
                title = (item.findtext('title') or '').strip()
                link  = (item.findtext('link') or '').strip()
                pub   = (item.findtext('pubDate') or '').strip()
                src_el = item.find('source')
                source = (
                    (src_el.text.strip() if src_el is not None and src_el.text else None)
                    or default_src or 'News'
                )
                if title and link and link.startswith('https://'):
                    items.append({'title': title, 'link': link, 'pubDate': pub, 'source': source})
        except Exception as exc:
            print(f'  ⚠  RSS parse error: {exc}')
    seen, out = set(), []
    for it in items:
        if it['title'] not in seen:
            seen.add(it['title'])
            out.append(it)
        if len(out) >= 10:
            break
    print(f'  {len(out)} news items')
    return out or None

# ══════════════════════════════════════════════════════════════════════════════
#  TICKER BAR
# ══════════════════════════════════════════════════════════════════════════════

def fetch_tickers():
    """Fetch latest price + % change for ticker bar symbols."""
    syms = list(TICKER_SYMBOLS.values())
    try:
        raw = yf.download(syms, period='2d', interval='1d', progress=False,
                          auto_adjust=True, threads=True)
    except Exception as exc:
        print(f'  ✗  Ticker download failed: {exc}')
        return None
    result = {}
    for key, sym in TICKER_SYMBOLS.items():
        try:
            closes = raw['Close'][sym].dropna() if sym in raw['Close'].columns else raw['Close'].dropna()
            if len(closes) < 1:
                continue
            price = float(closes.iloc[-1])
            prev  = float(closes.iloc[-2]) if len(closes) >= 2 else price
            pct   = (price - prev) / prev * 100 if prev else 0.0
            result[key] = {'price': round(price, 4), 'pct': round(pct, 4)}
        except Exception:
            continue
    print(f'  {len(result)} tickers')
    return result or None

# ══════════════════════════════════════════════════════════════════════════════
#  GAINERS  (uses dynamic universe)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_gainers(ex_key, tickers, names):
    """Return top-10 % gainers for the given exchange universe."""
    cfg      = EXCHANGES[ex_key]
    is_pence = cfg['pence']
    currency = cfg['currency']
    try:
        raw = yf.download(tickers, period='2d', interval='1d', progress=False,
                          auto_adjust=True, threads=True)
    except Exception as exc:
        print(f'  ✗  {ex_key} gainers download: {exc}')
        return None

    is_multi = isinstance(raw.columns, pd.MultiIndex)
    if is_multi:
        if 'Close' not in raw.columns.get_level_values(0):
            return None
        close_df = raw['Close']
    else:
        if 'Close' not in raw.columns:
            return None
        close_df = raw[['Close']].rename(columns={'Close': tickers[0]})

    results = []
    for sym in tickers:
        try:
            closes = close_df[sym].dropna() if (is_multi and sym in close_df.columns) else close_df.dropna()
            if len(closes) < 2:
                continue
            curr = float(closes.iloc[-1])
            prev = float(closes.iloc[-2])
            if prev <= 0:
                continue
            pct = (curr - prev) / prev * 100
            results.append({'sym': sym, 'price': curr, 'chg': curr - prev, 'pct': pct})
        except Exception:
            continue

    if not results:
        return None
    results.sort(key=lambda x: x['pct'], reverse=True)

    suffix_map = {'lse': '.L', 'enx': None, 'ndx': None, 'nyse': None, 'xetra': '.DE', 'six': '.SW', 'tse': '.T'}
    suffix = suffix_map.get(ex_key)
    output = []
    for r in results[:10]:
        sym = r['sym']
        p, chg, pct = r['price'], r['chg'], r['pct']
        display      = names.get(sym, sym.split('.')[0] if suffix else sym)
        ticker_clean = sym.replace(suffix, '') if suffix and sym.endswith(suffix) else sym.split('.')[0]
        if is_pence:
            price_str = f"{int(p)}p" if p >= 100 else f"{p:.2f}p"
            chg_str   = f"{'+' if chg>=0 else ''}{int(chg)}p" if abs(chg) >= 1 else f"{'+' if chg>=0 else ''}{chg:.2f}p"
        else:
            price_str = f"{currency}{p:.2f}"
            chg_str   = f"{'+' if chg>=0 else ''}{currency}{chg:.2f}"
        output.append({'name': display, 'ticker': ticker_clean,
                       'price': price_str, 'change': chg_str, 'pct': f"{pct:.2f}%"})

    print(f'  {len(output)} gainers · top: {output[0]["name"]} +{output[0]["pct"]}')
    return output

# ══════════════════════════════════════════════════════════════════════════════
#  MARKET CAP  (uses dynamic universe)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_market_cap(ex_key, tickers, names):
    """Return top-10 stocks by market cap from the dynamic universe."""
    cfg        = EXCHANGES[ex_key]
    is_pence   = cfg['pence']
    currency   = cfg['currency']
    mc_divisor = 1e11 if is_pence else 1e9

    try:
        tickers_obj = yf.Tickers(' '.join(tickers))
    except Exception as exc:
        print(f'  ✗  {ex_key} market cap: {exc}')
        return None

    results = []
    for sym in tickers:
        try:
            fi = tickers_obj.tickers[sym].fast_info
            mc    = getattr(fi, 'market_cap', None) or getattr(fi, 'marketCap', None)
            price = getattr(fi, 'last_price',  None) or getattr(fi, 'lastPrice',  0) or 0
            if mc and float(mc) > 0:
                results.append({'sym': sym, 'mc': float(mc), 'price': float(price)})
        except Exception:
            continue

    if not results:
        return None
    results.sort(key=lambda x: x['mc'], reverse=True)

    suffix_map = {'lse': '.L', 'enx': None, 'ndx': None, 'nyse': None, 'xetra': '.DE', 'six': '.SW', 'tse': '.T'}
    suffix = suffix_map.get(ex_key)
    top10  = []
    for r in results[:10]:
        sym  = r['sym']
        mc_b = r['mc'] / mc_divisor
        p    = r['price']
        display      = names.get(sym, sym.split('.')[0] if suffix else sym)
        ticker_clean = sym.replace(suffix, '') if suffix and sym.endswith(suffix) else sym.split('.')[0]
        price_str    = f'{int(p)}p' if is_pence and p >= 100 else (f'{p:.2f}p' if is_pence else f'{currency}{p:.2f}')
        top10.append({'name': display, 'ticker': ticker_clean,
                      'mcap_b': round(mc_b, 1), 'price': price_str})

    print(f'  {len(top10)} stocks · largest: {top10[0]["name"]} {top10[0]["mcap_b"]:.1f}B')
    return {'date': _today(), 'currency': currency, 'top10': top10}

# ══════════════════════════════════════════════════════════════════════════════
#  VOLUME — file-based (LSE and Euronext official publications)
# ══════════════════════════════════════════════════════════════════════════════

LSE_URL = ('https://docs.londonstockexchange.com/sites/default/files'
           '/reports/Order%20book%20trading_{num}.xlsx')

def _lse_fetch(state):
    last_num = state.get('lse_file_num', 1513)
    latest   = last_num
    for n in range(last_num + 1, last_num + 40):
        try:
            req = urllib.request.Request(LSE_URL.format(num=n), method='HEAD', headers=_HDR)
            with urllib.request.urlopen(req, context=_SSL, timeout=10) as r:
                if r.status == 200:
                    latest = n
                else:
                    break
        except Exception:
            break
    raw = _get(LSE_URL.format(num=latest))
    if raw:
        state['lse_file_num'] = latest
    return raw

def _lse_parse(raw):
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb['Daily Order Book Trading']
    rows, header = [], False
    for row in ws.iter_rows(values_only=True):
        if row[1] == 'Trade Date':
            header = True
            continue
        if header and isinstance(row[1], datetime) and row[4] is not None:
            rows.append({
                'label':  row[1].strftime('%Y-%m-%d'),
                'value':  round(float(row[5]) / 1e9, 2),
                'trades': round(int(row[4]) / 1000, 1),
            })
        if len(rows) >= 5:
            break
    return list(reversed(rows[:5]))

def _enx_url(dt):
    d  = dt.strftime('%Y%m%d')
    yr = dt.strftime('%Y')
    base = 'https://live.euronext.com/sites/default/files/statistics/cash/nextday'
    return f'{base}/{yr}/Cash%20{d}.xlsx', f'{base}/2017/Cash%20{d}.xlsx'

def _enx_fetch(dt):
    url_yr, url_17 = _enx_url(dt)
    raw = _get(url_yr)
    if raw is None:
        raw = _get(url_17)
    return raw

def _enx_parse(raw):
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb['All Markets']
    rows = list(ws.iter_rows(values_only=True))
    date_row     = rows[4]
    turnover_row = rows[6]
    trades_row   = rows[22] if len(rows) > 22 else None
    pts = []
    for col in [4, 3, 2]:
        d = date_row[col]
        v = turnover_row[col]
        if isinstance(d, datetime) and v is not None:
            pt = {'label': d.strftime('%Y-%m-%d'), 'value': round(float(v) / 1000, 2)}
            if trades_row is not None:
                try:
                    t = trades_row[col]
                    if t is not None:
                        pt['trades'] = round(float(t) / 1_000_000, 2)
                except Exception:
                    pass
            pts.append(pt)
    return pts

def fetch_vol_file(ex_key, state):
    """Download and parse LSE or Euronext official volume file.
    Returns list of {'label', 'value'} dicts (oldest→newest, up to 5 days).
    """
    if ex_key == 'lse':
        raw = _lse_fetch(state)
        if raw is None:
            return None
        return _lse_parse(raw)

    today_dt, today_raw = date.today(), None
    for _ in range(4):
        today_raw = _enx_fetch(today_dt)
        if today_raw:
            break
        today_dt = _prev_bday(today_dt)
    if today_raw is None:
        print('  ✗  Euronext: could not fetch any file')
        return None

    today_pts = _enx_parse(today_raw)
    if len(today_pts) < 2:
        return today_pts or None

    try:
        oldest_label = today_pts[0]['label']
        t2_dt = date.fromisoformat(oldest_label) - timedelta(days=3)
        # Walk back to a business day
        while t2_dt.weekday() >= 5:
            t2_dt -= timedelta(days=1)
    except Exception:
        return today_pts

    t2_raw = _enx_fetch(t2_dt)
    if t2_raw is None:
        return today_pts

    t2_pts = _enx_parse(t2_raw)
    combined = t2_pts[:2] + today_pts
    return combined or None

# ──────────────────────────────────────────────────────────────────────────────
#  VOLUME — yfinance universe estimate (NDX / NYSE / Xetra / SIX)
# ──────────────────────────────────────────────────────────────────────────────

def fetch_vol_yf(ex_key, tickers):
    """Compute aggregate daily turnover from the dynamic universe tickers."""
    cfg = EXCHANGES[ex_key]
    try:
        raw = yf.download(tickers, period='15d', interval='1d',
                          progress=False, auto_adjust=True, threads=True)
        if raw.empty:
            print(f'  ✗  {ex_key} vol: no data')
            return None
    except Exception as exc:
        print(f'  ✗  {ex_key} vol download: {exc}')
        return None

    try:
        close_df  = raw['Close']  if isinstance(raw.columns, pd.MultiIndex) else raw
        volume_df = raw['Volume'] if isinstance(raw.columns, pd.MultiIndex) else None
        if volume_df is None:
            return None
        daily_turnover = (close_df * volume_df).sum(axis=1, min_count=1) / 1e9
        daily_turnover = daily_turnover.dropna()
        daily_shares   = volume_df.sum(axis=1, min_count=1) / 1e6
        pts = []
        for ts, val in daily_turnover.items():
            label = ts.strftime('%Y-%m-%d')
            if val > 0:
                sh = float(daily_shares.get(ts, 0)) if ts in daily_shares.index else 0.0
                pts.append({'label': label, 'value': round(float(val), 2),
                            'shares_m': round(sh, 1) if sh > 0 else None})
        pts = pts[-5:]
        print(f'  {len(pts)} vol days · last: {pts[-1]["value"]:.1f}B ({ex_key})' if pts else f'  ✗  {ex_key}: no vol points')
        return pts or None
    except Exception as exc:
        print(f'  ✗  {ex_key} vol parse: {exc}')
        return None

# ──────────────────────────────────────────────────────────────────────────────
#  COMPARABLE VOLUME  (all 7 exchanges — same liquid universe)
# ──────────────────────────────────────────────────────────────────────────────

def fetch_vol_comparable(ex_key, tickers, fx_rates):
    """Compute daily aggregate turnover for the liquid universe tickers.
    Returns list of {'date', 'value_local', 'value_usd'} for the last 5 trading days.
    All six exchanges go through this path so volume rankings compare like with like.
    """
    cfg      = EXCHANGES[ex_key]
    is_pence = cfg.get('pence', False)
    nok_eur  = cfg.get('nok_eur', False)
    currency = cfg['currency']
    fx_usd   = fx_rates.get(currency, 1.0)

    try:
        raw = yf.download(tickers, period='15d', interval='1d',
                          progress=False, auto_adjust=True, threads=True)
        if raw.empty:
            print(f'  ✗  {ex_key} vol_comparable: no data')
            return None
    except Exception as exc:
        print(f'  ✗  {ex_key} vol_comparable download: {exc}')
        return None

    is_multi = isinstance(raw.columns, pd.MultiIndex)
    if not is_multi:
        return None

    close_df  = raw['Close']
    volume_df = raw['Volume']
    daily_val = close_df * volume_df   # local_ccy · shares

    if is_pence:
        daily_val = daily_val / 100    # pence → GBP

    if nok_eur:
        nok_eur_rate = fx_rates.get('NOK', 0.086) / fx_rates.get('EUR', 1.08)
        ol_cols = [t for t in tickers if t.endswith('.OL') and t in daily_val.columns]
        for t in ol_cols:
            daily_val[t] = daily_val[t] * nok_eur_rate   # NOK → EUR

    daily_total_local  = daily_val.sum(axis=1, min_count=1).dropna()
    daily_shares_total = volume_df.sum(axis=1, min_count=1)   # total shares traded, universe-wide

    pts = []
    for ts, val_local in daily_total_local.items():
        if val_local <= 0:
            continue
        val_usd = float(val_local) * fx_usd
        sh = float(daily_shares_total.get(ts, 0)) if ts in daily_shares_total.index else 0.0
        pts.append({
            'date':        ts.strftime('%Y-%m-%d'),
            'value_local': round(float(val_local) / 1e9, 3),   # → local currency billions
            'value_usd':   round(val_usd / 1e9, 3),            # → USD billions
            'shares_m':    round(sh / 1e6, 1) if sh > 0 else None,   # → millions of shares
        })

    pts = pts[-5:]

    # Drop today's point while global markets are still trading. The yfinance
    # 'today' bar during market hours only reflects volume so far, not the
    # full session — including it would skew the Latest/5-day-avg figures
    # shown here, and the vol sub-rank in the Liquidity Ranking (which reads
    # this same series). Same coarse UTC-hour cutoff used in compute_rankings.
    today_iso = date.today().isoformat()
    if pts and pts[-1]['date'] == today_iso and datetime.now(timezone.utc).hour < 20:
        pts = pts[:-1]

    if pts:
        print(f'  vol_comparable {ex_key}: {pts[-1]["value_local"]:.2f}B local / '
              f'{pts[-1]["value_usd"]:.2f}B USD ({len(tickers)} stocks)')
    return pts or None

# ══════════════════════════════════════════════════════════════════════════════
#  AMIHUD ILLIQUIDITY  (uses dynamic universe)
#  ILLIQ = |R| / DVOL_M  (% price move per 1M currency unit traded)
# ══════════════════════════════════════════════════════════════════════════════

def _fetch_nok_eur():
    try:
        raw = yf.download('NOKEUR=X', period='2d', interval='1d',
                          progress=False, auto_adjust=True)
        return float(raw['Close'].dropna().iloc[-1])
    except Exception:
        return 0.086

def fetch_amihud(ex_key, tickers, names, existing_history):
    """Compute per-day Amihud ILLIQ for the exchange universe.
    Returns (updated_history, marketAvg, top10_liquid, top10_illiquid, top10_active, top10_inactive).
    """
    cfg      = EXCHANGES[ex_key]
    is_pence = cfg['pence']
    nok_eur  = _fetch_nok_eur() if cfg['nok_eur'] else None

    try:
        raw = yf.download(tickers, period='40d', interval='1d', progress=False,
                          auto_adjust=True, threads=True)
        if raw.empty:
            print(f'  ✗  {ex_key} amihud: no data')
            return existing_history, None, [], [], [], []
    except Exception as exc:
        print(f'  ✗  {ex_key} amihud download: {exc}')
        return existing_history, None, [], [], [], []

    daily_buckets = {}
    stock_avgs    = {}
    dvol_avgs     = {}

    for sym in tickers:
        try:
            df = raw[[('Close', sym), ('Volume', sym)]].copy()
            df.columns = ['Close', 'Volume']
            df = df.dropna()
            if len(df) < 6:
                continue
            if is_pence:
                dvol_m = df['Close'] * df['Volume'] / 100 / 1_000_000
            else:
                dvol_m = df['Close'] * df['Volume'] / 1_000_000
                if cfg['nok_eur'] and sym.endswith('.OL') and nok_eur:
                    dvol_m = dvol_m * nok_eur
            ret_pct = df['Close'].pct_change().abs() * 100
            illiq   = (ret_pct / dvol_m).replace([float('inf'), float('-inf')], float('nan')).dropna()
            illiq   = illiq.iloc[-20:]
            if len(illiq) < 5:
                continue
            stock_avgs[sym] = float(illiq.mean())
            dvol_avgs[sym]  = float(dvol_m.iloc[-20:].mean())
            for ts, val in illiq.items():
                if pd.isna(val):
                    continue
                label = ts.strftime('%Y-%m-%d')
                daily_buckets.setdefault(label, []).append(float(val))
        except Exception:
            continue

    if not stock_avgs:
        print(f'  ✗  {ex_key}: no ILLIQ computed')
        return existing_history, None, [], [], [], []

    market_avg = sum(stock_avgs.values()) / len(stock_avgs)

    try:
        last_trade_date = raw.dropna(how='all').index[-1].strftime('%Y-%m-%d')
    except Exception:
        last_trade_date = None

    bucket_dates = sorted(daily_buckets.keys())
    print(f'  bucket dates (last 3): {bucket_dates[-3:]} · last_trade: {last_trade_date} ({ex_key})')

    new_history = list(existing_history)
    for label in sorted(daily_buckets):
        vals = daily_buckets[label]
        new_history = _push(new_history, {'date': label, 'illiq': round(sum(vals)/len(vals), 6)})

    if (last_trade_date and
            not any(e.get('date') == last_trade_date for e in new_history)):
        print(f'  ⚠  {ex_key}: {last_trade_date} absent — injecting market_avg as fallback')
        new_history = _push(new_history, {'date': last_trade_date, 'illiq': round(market_avg, 6)})

    sorted_stocks = sorted(stock_avgs.items(), key=lambda x: x[1])

    suffix_map = {'lse':'.L','enx':None,'ndx':None,'nyse':None,'xetra':'.DE','six':'.SW','tse':'.T'}
    suf = suffix_map.get(ex_key)

    def _stock_entry(sym, val):
        tk = sym.replace(suf, '') if suf and sym.endswith(suf) else sym.split('.')[0]
        return {'ticker': tk, 'name': names.get(sym, tk), 'illiq': round(val, 6)}

    def _dvol_entry(sym, val):
        tk = sym.replace(suf, '') if suf and sym.endswith(suf) else sym.split('.')[0]
        return {'ticker': tk, 'name': names.get(sym, tk), 'dvol_m': round(val, 2)}

    top10_liquid   = [_stock_entry(s, v) for s, v in sorted_stocks[:10]]
    top10_illiquid = [_stock_entry(s, v) for s, v in reversed(sorted_stocks[-10:])]

    sorted_dvol    = sorted(dvol_avgs.items(), key=lambda x: x[1])
    top10_active   = [_dvol_entry(s, v) for s, v in reversed(sorted_dvol[-10:])]
    top10_inactive = [_dvol_entry(s, v) for s, v in sorted_dvol[:10]]

    print(f'  {len(stock_avgs)} stocks · market avg ILLIQ: {market_avg:.4f} ({ex_key})')
    return new_history, round(market_avg, 6), top10_liquid, top10_illiquid, top10_active, top10_inactive

# ══════════════════════════════════════════════════════════════════════════════
#  ABDI-RANALDO (2017) IMPLIED SPREAD  (uses dynamic universe)
#  Replaced the Roll (1984) implied spread — AR is more robust for liquid
#  large-caps since it uses intraday High/Low, not just close-to-close.
# ══════════════════════════════════════════════════════════════════════════════

def _ar_spread(highs, lows, closes):
    data = [(float(c), float(h), float(l))
            for c, h, l in zip(closes, highs, lows)
            if c and h and l and float(c) > 0 and float(h) > 0 and float(l) > 0]
    if len(data) < 12:
        return None
    ct = [math.log(c) - 0.5 * (math.log(h) + math.log(l)) for c, h, l in data]
    if len(ct) < 8:
        return None
    c1, c2 = ct[:-1], ct[1:]
    m1 = sum(c1) / len(c1)
    m2 = sum(c2) / len(c2)
    cov = sum((a - m1) * (b - m2) for a, b in zip(c1, c2)) / (len(c1) - 1)
    if cov >= 0:
        return None
    s = round(2 * ((-cov) ** 0.5) * 100, 4)
    return s if 0 < s <= 5 else None

def fetch_ar_spread(ex_key, tickers, names, existing_history):
    """Compute Abdi-Ranaldo spread for exchange universe.
    Returns (updated_history, marketAvg, top10_tight, top10_wide).
    """
    try:
        raw = yf.download(tickers, period='60d', interval='1d', progress=False,
                          auto_adjust=True)
        if raw.empty:
            return existing_history, None, [], []
    except Exception as exc:
        print(f'  ✗  {ex_key} AR download: {exc}')
        return existing_history, None, [], []

    is_multi = isinstance(raw.columns, pd.MultiIndex)
    try:
        close_df = raw['Close']  if is_multi else raw
        high_df  = raw['High']   if is_multi else None
        low_df   = raw['Low']    if is_multi else None
        if high_df is None or low_df is None:
            return existing_history, None, [], []
        last_trade_date = close_df.dropna(how='all').index[-1].strftime('%Y-%m-%d')
    except Exception:
        last_trade_date = _today()

    stock_spreads = {}
    for sym in tickers:
        try:
            if sym not in close_df.columns:
                continue
            cl  = close_df[sym].dropna()
            hi  = high_df[sym].reindex(cl.index).dropna()
            lo  = low_df[sym].reindex(cl.index).dropna()
            idx = cl.index.intersection(hi.index).intersection(lo.index)
            sp  = _ar_spread(hi.loc[idx].tolist(), lo.loc[idx].tolist(), cl.loc[idx].tolist())
            if sp is not None:
                stock_spreads[sym] = sp
        except Exception:
            continue

    if not stock_spreads:
        print(f'  ✗  {ex_key}: no AR spread computed')
        return existing_history, None, [], []

    market_avg  = round(sum(stock_spreads.values()) / len(stock_spreads), 4)
    new_history = _push(existing_history, {'date': last_trade_date, 'avgSpread': market_avg})

    suffix_map = {'lse':'.L','enx':None,'ndx':None,'nyse':None,'xetra':'.DE','six':'.SW','tse':'.T'}
    suf = suffix_map.get(ex_key)

    def _ar_entry(sym, val):
        tk = sym.replace(suf, '') if suf and sym.endswith(suf) else sym.split('.')[0]
        return {'ticker': tk, 'name': names.get(sym, tk), 'spread': round(val, 4)}

    sorted_s    = sorted(stock_spreads.items(), key=lambda x: x[1])
    top10_tight = [_ar_entry(s, v) for s, v in sorted_s[:10]]
    top10_wide  = [_ar_entry(s, v) for s, v in reversed(sorted_s[-10:])]

    print(f'  {len(stock_spreads)} stocks · AR spread: {market_avg:.4f}% ({ex_key}, {last_trade_date})')
    return new_history, market_avg, top10_tight, top10_wide

# ══════════════════════════════════════════════════════════════════════════════
#  RANKING COMPUTATION
#  Uses vol_comparable (USD) for volume sub-rank so all exchanges are comparable.
# ══════════════════════════════════════════════════════════════════════════════

def _rank6(vals_dict, ascending=True):
    """Rank exchange values 1–N. ascending=True → lowest value = rank 1."""
    items = [(k, v) for k, v in vals_dict.items() if v is not None]
    items.sort(key=lambda x: x[1], reverse=not ascending)
    return {k: i+1 for i, (k, _) in enumerate(items)}

def compute_rankings(dash):
    """Build current_ranking and ranking_history from stored metric histories.
    Volume sub-rank uses vol_comparable (USD) for cross-exchange comparability.
    """
    ex_keys     = list(EXCHANGES.keys())
    today_label = _today()

    _utc_hour    = datetime.now(timezone.utc).hour
    _exclude_today = today_label if _utc_hour < 20 else '__never_match__'

    illiq_by_ex = {}
    for k in ex_keys:
        illiq_by_ex[k] = {e['date']: e['illiq']
                          for e in dash['amihud'][k].get('history', [])
                          if e.get('date') != _exclude_today}

    # Volume sub-rank: use vol_comparable value_usd (USD billions, same for all)
    vol_latest = {}
    for k in ex_keys:
        vc = dash.get('vol_comparable', {}).get(k, {})
        vals = vc.get('value_usd', [])
        # Fallback to vol if comparable not yet available
        if not vals:
            vd = dash['vol'][k]
            vals = vd.get('value', [])
        vol_latest[k] = vals[-1] if vals else None

    tr_by_ex = {}
    for k in ex_keys:
        tr_by_ex[k] = {e['date']: e['ratio']
                       for e in dash.get('turnover_ratio', {}).get(k, {}).get('history', [])
                       if e.get('date') != _exclude_today}

    ar_by_ex = {}
    for k in ex_keys:
        ar_by_ex[k] = {e['date']: e['avgSpread']
                       for e in dash['ar_spread'][k].get('history', [])
                       if e.get('date') != _exclude_today}

    for k in ex_keys:
        dates = sorted(illiq_by_ex[k].keys())
        print(f'  ILLIQ {k}: {len(dates)} dates · latest 3: {dates[-3:] if len(dates) >= 3 else dates}')

    all_dates = set()
    for k in ex_keys:
        all_dates.update(illiq_by_ex[k].keys())

    min_open = max(3, len(ex_keys) - 2)
    complete_dates = sorted(
        [d for d in all_dates
         if sum(1 for k in ex_keys if d in illiq_by_ex[k]) >= min_open]
    )

    if not complete_dates:
        print('  ⚠  Rankings: no complete dates found')
        return

    def _best_on_or_before(lookup, d):
        v = lookup.get(d)
        if v is not None:
            return v
        prior = sorted(dt for dt in lookup if dt <= d)
        return lookup[prior[-1]] if prior else None

    history = []
    for d in complete_dates:
        illiq_vals = {k: _best_on_or_before(illiq_by_ex[k], d) for k in ex_keys}
        vol_vals   = {k: vol_latest[k]                           for k in ex_keys}
        tr_vals    = {k: _best_on_or_before(tr_by_ex[k],    d) for k in ex_keys}
        ar_vals    = {k: _best_on_or_before(ar_by_ex[k],    d) for k in ex_keys}

        r_illiq = _rank6(illiq_vals, ascending=True)
        r_tr    = _rank6(tr_vals,    ascending=False)
        r_ar    = _rank6(ar_vals,    ascending=True)
        r_vol   = _rank6(vol_vals,   ascending=False)

        composites = {}
        for k in ex_keys:
            sub   = [r_illiq.get(k), r_tr.get(k), r_ar.get(k), r_vol.get(k)]
            avail = [x for x in sub if x is not None]
            if avail:
                composites[k] = sum(avail) / len(avail)
        r_comp = _rank6(composites, ascending=True)

        ranks = {k: {
            'vol':       r_vol.get(k),
            'illiq':     r_illiq.get(k),
            'tr':        r_tr.get(k),
            'ar':        r_ar.get(k),
            'composite': r_comp.get(k),
        } for k in ex_keys}
        history.append({'date': d, 'ranks': ranks})

    dash['ranking_history'] = history[-DAYS:]
    last = history[-1]
    dash['current_ranking'] = {'date': last['date'], 'ranks': last['ranks']}
    print(f'  Ranking computed · {len(complete_dates)} complete days · current: {last["date"]}')
    sorted_ex = sorted(ex_keys, key=lambda k: last['ranks'][k].get('composite') or 99)
    for rank, k in enumerate(sorted_ex, 1):
        print(f'    #{rank} {EXCHANGES[k]["name"]} (composite rank {last["ranks"][k].get("composite")})')

# ══════════════════════════════════════════════════════════════════════════════
#  AI COMMENTARY
# ══════════════════════════════════════════════════════════════════════════════

def generate_commentary(dash):
    """Generate a short AI ranking commentary. Returns string or None."""
    if _ant is None:
        print('  ⚠  anthropic not installed — commentary skipped')
        return None
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        print('  ⚠  ANTHROPIC_API_KEY not set — commentary skipped')
        return None

    cr = dash.get('current_ranking', {})
    ranking_date = cr.get('date', 'unknown date')
    ranks = cr.get('ranks', {})
    if not ranks:
        print('  ⚠  No ranking data for commentary')
        return None

    sorted_ex = sorted(EXCHANGES.keys(),
                       key=lambda k: ranks.get(k, {}).get('composite') or 99)

    def amihud_val(k):
        avg = dash['amihud'][k].get('marketAvg')
        return f'{avg:.4f}' if avg is not None else 'N/A'
    def ar_val(k):
        avg = dash['ar_spread'][k].get('marketAvg')
        return f'{avg:.4f}%' if avg is not None else 'N/A'

    ranking_summary = '\n'.join(
        f'  #{i+1} {EXCHANGES[k]["name"]}: composite rank {ranks[k].get("composite")}, '
        f'ILLIQ={amihud_val(k)}, AR spread={ar_val(k)}'
        for i, k in enumerate(sorted_ex)
    )

    prompt = f"""You are a financial markets analyst writing a brief commentary on exchange liquidity rankings.

Today's data is for {ranking_date}.

Exchange Liquidity Ranking (1 = most liquid):
{ranking_summary}

Metrics explained:
- ILLIQ (Amihud): % price move per 1M currency unit traded — lower = more liquid
- AR spread: Abdi-Ranaldo (2017) implied spread — lower = tighter spreads
- Volume: comparable liquid-universe turnover (same constituent methodology, USD) — higher = more liquid
- Turnover ratio: comparable volume / market cap — higher = more liquid

Write a concise 3-4 sentence commentary (max 80 words) for a professional audience. Focus on the top and bottom ranked exchanges and any notable differences in the metrics. Be specific and analytical. Do not use bullet points."""

    try:
        client = _ant.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=200,
            messages=[{'role': 'user', 'content': prompt}]
        )
        text = msg.content[0].text.strip()
        print(f'  ✓  Commentary generated ({len(text)} chars)')
        return text
    except Exception as exc:
        print(f'  ✗  Commentary generation failed: {exc}')
        return None

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Finance Dashboard Updater')
    parser.add_argument('--fast', action='store_true',
                        help='Fast mode: update ticker bar, gainers, and market cap only')
    args     = parser.parse_args()
    FAST_MODE = args.fast

    today_iso = date.today().isoformat()

    # ── State (lives inside dashboard_data.json under '_state') ──────────────
    dash_for_state = _load_dashboard()
    state = dash_for_state.pop('_state', {})
    if not state and STATE_FILE.exists():
        try:
            state = json.loads(STATE_FILE.read_text())
        except Exception:
            pass

    mode_label = 'FAST (news / ticker / gainers / market cap)' if FAST_MODE else 'FULL'
    print('\n═══ Finance Dashboard Updater ═══')
    print(f'Date: {date.today()} UTC   Mode: {mode_label}\n')

    dash = _load_dashboard()
    _migrate_dates(dash)

    # ── Methodology version check — clear histories on universe change ────────
    if not FAST_MODE and dash.get('methodology_version') != METHODOLOGY_VERSION:
        print('  ℹ  Methodology version change detected — clearing metric histories')
        for section in ('amihud', 'spread', 'ar_spread', 'turnover_ratio'):
            for k in EXCHANGES:
                if section in dash and k in dash[section]:
                    dash[section][k]['history'] = []
                    dash[section][k].pop('marketAvg', None)
        dash['ranking_history'] = []
        dash['current_ranking'] = {'date': None, 'ranks': {
            k: {'vol': None, 'illiq': None, 'tr': None, 'ar': None, 'composite': None}
            for k in EXCHANGES
        }}
        dash['methodology_version'] = METHODOLOGY_VERSION
        print('  ✓  Histories cleared — fresh start with dynamic universe\n')

    # ── News ─────────────────────────────────────────────────────────────────
    print('── News')
    news = fetch_news()
    if news:
        dash['news'] = news

    # ── Ticker bar ───────────────────────────────────────────────────────────
    print('\n── Tickers')
    tickers_data = fetch_tickers()
    if tickers_data:
        dash['tickers'] = tickers_data

    if not FAST_MODE:
        # ── Universe refresh (monthly) ────────────────────────────────────────
        print('\n── Universe')
        last_refresh  = state.get('universe_refresh_date', '')
        no_universe   = not any(
            dash.get('universes', {}).get(k, {}).get('tickers')
            for k in EXCHANGES
        )
        needs_refresh = no_universe or (
            last_refresh and
            (date.today() - date.fromisoformat(last_refresh)).days >= 30
        ) or (not last_refresh)

        if needs_refresh:
            print('  Refreshing universes from Wikipedia...')
            fx_rates = _fetch_fx_rates()
            if 'universes' not in dash:
                dash['universes'] = {}
            built = 0
            for ex_key in EXCHANGES:
                result = build_universe(ex_key, fx_rates)
                if result:
                    dash['universes'][ex_key] = result
                    built += 1
                else:
                    print(f'  ⚠  {ex_key}: keeping cached universe')
            if built > 0:
                state['universe_refresh_date'] = today_iso
                print(f'  ✓  Universe refresh complete ({built}/{len(EXCHANGES)} exchanges updated)')
        else:
            days_since = (date.today() - date.fromisoformat(last_refresh)).days
            print(f'  ℹ  Universe current (refreshed {days_since}d ago, next refresh in {30-days_since}d)')
            fx_rates = _fetch_fx_rates()   # still need FX for vol_comparable
    else:
        fx_rates = {}

    # ── Per-exchange data ─────────────────────────────────────────────────────
    for ex_key, cfg in EXCHANGES.items():
        print(f'\n── {cfg["name"]} ({ex_key})')

        tickers, names = _get_universe(dash, ex_key)

        # Gainers
        print('   Gainers...')
        gainers = fetch_gainers(ex_key, tickers, names)
        if gainers:
            dash['gainers'][ex_key] = gainers

        # Market Cap
        print('   Market Cap...')
        mcap = fetch_market_cap(ex_key, tickers, names)
        if mcap:
            dash['market_cap'][ex_key] = mcap

        if FAST_MODE:
            continue

        # Official Volume (LSE and ENX from publications; others from yfinance universe)
        print('   Volume (official/universe)...')
        if cfg['vol_method'] == 'file':
            vol_pts = fetch_vol_file(ex_key, state)
        else:
            vol_pts = fetch_vol_yf(ex_key, tickers)
        if vol_pts:
            existing_dates = list(dash['vol'][ex_key].get('dates', []))
            existing_vals  = list(dash['vol'][ex_key].get('value', []))
            vol_map = dict(zip(existing_dates, existing_vals))
            for pt in vol_pts:
                vol_map[pt['label']] = pt['value']
            sorted_labels = _trim_stale_dates(vol_map.keys())
            dash['vol'][ex_key]['dates'] = sorted_labels
            dash['vol'][ex_key]['value'] = [vol_map[l] for l in sorted_labels]
            valid_sh = [pt['shares_m'] for pt in vol_pts if pt.get('shares_m') is not None]
            if valid_sh:
                dash['vol'][ex_key]['shares_latest'] = valid_sh[-1]
                dash['vol'][ex_key]['shares_avg']    = round(sum(valid_sh) / len(valid_sh), 1)
            valid_tr = [pt['trades'] for pt in vol_pts if pt.get('trades') is not None]
            if valid_tr:
                dash['vol'][ex_key]['trades_latest'] = valid_tr[-1]
                dash['vol'][ex_key]['trades_avg']    = round(sum(valid_tr) / len(valid_tr), 1)

        # Comparable Volume (same liquid universe for all exchanges)
        print('   Volume (comparable universe)...')
        if 'vol_comparable' not in dash:
            dash['vol_comparable'] = {k: {
                'currency': EXCHANGES[k]['vol_currency'], 'currency_usd': 'USD B',
                'dates': [], 'value': [], 'value_usd': [], 'shares': [], 'n_universe': 0,
            } for k in EXCHANGES}
        vc_pts = fetch_vol_comparable(ex_key, tickers, fx_rates)
        if vc_pts:
            vc = dash['vol_comparable'][ex_key]
            # Merge into existing comparable vol history (keep last 5)
            vc_map_local  = dict(zip(vc.get('dates', []), vc.get('value', [])))
            vc_map_usd    = dict(zip(vc.get('dates', []), vc.get('value_usd', [])))
            vc_map_shares = dict(zip(vc.get('dates', []), vc.get('shares', [])))
            for pt in vc_pts:
                vc_map_local[pt['date']] = pt['value_local']
                vc_map_usd[pt['date']]   = pt['value_usd']
                if pt.get('shares_m') is not None:
                    vc_map_shares[pt['date']] = pt['shares_m']
            sorted_vc_dates = _trim_stale_dates(vc_map_local.keys())
            vc['dates']      = sorted_vc_dates
            vc['value']      = [vc_map_local[d] for d in sorted_vc_dates]
            vc['value_usd']  = [vc_map_usd[d]   for d in sorted_vc_dates]
            vc['shares']     = [vc_map_shares.get(d) for d in sorted_vc_dates]
            vc['n_universe'] = len(tickers)
            valid_shares = [s for s in vc['shares'] if s is not None]
            if valid_shares:
                vc['shares_latest'] = valid_shares[-1]
                vc['shares_avg']    = round(sum(valid_shares) / len(valid_shares), 1)

        # Amihud ILLIQ
        print('   Amihud ILLIQ...')
        new_hist, mkt_avg, top_liq, top_illiq, top_active, top_inactive = fetch_amihud(
            ex_key, tickers, names, dash['amihud'][ex_key].get('history', []))
        dash['amihud'][ex_key]['history'] = new_hist
        if mkt_avg is not None:
            dash['amihud'][ex_key]['marketAvg']      = mkt_avg
            dash['amihud'][ex_key]['top10_liquid']   = top_liq
            dash['amihud'][ex_key]['top10_illiquid'] = top_illiq
        if 'turnover_ratio' not in dash:
            dash['turnover_ratio'] = {k: {'history': []} for k in EXCHANGES}
        dash['turnover_ratio'][ex_key]['top10_active']   = top_active
        dash['turnover_ratio'][ex_key]['top10_inactive'] = top_inactive

        # AR Spread
        print('   AR spread...')
        new_hist, mkt_avg, top_tight, top_wide = fetch_ar_spread(
            ex_key, tickers, names, dash['ar_spread'][ex_key].get('history', []))
        dash['ar_spread'][ex_key]['history'] = new_hist
        if mkt_avg is not None:
            dash['ar_spread'][ex_key]['marketAvg']   = mkt_avg
            dash['ar_spread'][ex_key]['top10_tight'] = top_tight
            dash['ar_spread'][ex_key]['top10_wide']  = top_wide

    if not FAST_MODE:
        # ── Turnover Ratios (uses vol_comparable as numerator) ────────────────
        print('\n── Turnover Ratios')
        if 'turnover_ratio' not in dash:
            dash['turnover_ratio'] = {k: {'history': []} for k in EXCHANGES}
        for ek in EXCHANGES:
            # Numerator: comparable volume (local currency billions)
            vc       = dash.get('vol_comparable', {}).get(ek, {})
            vc_dates = vc.get('dates', [])
            vc_vals  = vc.get('value', [])   # local currency B
            # Fallback to official vol if comparable not yet available
            if not vc_vals:
                vc_dates = dash['vol'][ek].get('dates', [])
                vc_vals  = dash['vol'][ek].get('value', [])
            # Denominator: market cap of universe top-10 (consistent with same universe)
            mcap_list  = dash['market_cap'][ek].get('top10', [])
            total_mcap = sum(t.get('mcap_b', 0) for t in mcap_list)
            if vc_vals and total_mcap > 0:
                for d, v in zip(vc_dates, vc_vals):
                    ratio = round(v / total_mcap * 100, 4)
                    dash['turnover_ratio'][ek]['history'] = _push(
                        dash['turnover_ratio'][ek].get('history', []),
                        {'date': d, 'ratio': ratio})
                latest = dash['turnover_ratio'][ek]['history'][-1]
                print(f'  {EXCHANGES[ek]["name"]}: {latest["ratio"]:.4f}% ({latest["date"]})')
            else:
                print(f'  {EXCHANGES[ek]["name"]}: insufficient data')

        # ── Rankings ──────────────────────────────────────────────────────────
        print('\n── Rankings')
        compute_rankings(dash)

        # ── AI Commentary ─────────────────────────────────────────────────────
        print('\n── Commentary')
        ranking_date = dash.get('current_ranking', {}).get('date')
        stale = (state.get('commentary_date') != today_iso or
                 state.get('commentary_ranking_date') != ranking_date or
                 not dash.get('commentary'))
        if not stale:
            print("  ℹ  Reusing today's commentary")
        else:
            commentary = generate_commentary(dash)
            if commentary:
                dash['commentary']               = commentary
                dash['commentary_date']          = today_iso
                state['commentary_date']         = today_iso
                state['commentary_ranking_date'] = ranking_date

    # ── Save ──────────────────────────────────────────────────────────────────
    dash['_state'] = state
    _save_dashboard(dash)
    print('  ✓  State saved (dashboard_data.json "_state")')
    print('\n═══ Done ═══\n')

if __name__ == '__main__':
    main()
